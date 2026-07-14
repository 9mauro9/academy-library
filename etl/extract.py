import openpyxl
import json
import os
import re
from datetime import time, timedelta

def slugify(text):
    if not text:
        return ""
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[-\s]+', '-', text)
    return text.strip('-')

def parse_duration(val):
    if val is None:
        return 0
    if isinstance(val, time):
        return val.hour * 3600 + val.minute * 60 + val.second
    if isinstance(val, timedelta):
        return int(val.total_seconds())
    if isinstance(val, (int, float)):
        return int(val)
    # try parsing string
    val_str = str(val).strip()
    if not val_str:
        return 0
    # Try parsing format HH:MM:SS
    parts = val_str.split(':')
    if len(parts) == 3:
        try:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
        except ValueError:
            pass
    elif len(parts) == 2:
        try:
            return int(parts[0]) * 60 + int(parts[1])
        except ValueError:
            pass
    try:
        return int(float(val_str))
    except ValueError:
        return 0

def clean_string(val):
    if val is None:
        return None
    val_str = str(val).strip()
    if val_str == "N/A" or val_str == "None" or not val_str:
        return None
    return val_str

def parse_tags(val):
    if not val:
        return []
    val_str = str(val).strip()
    if not val_str or val_str == "N/A":
        return []
    return [tag.strip() for tag in val_str.split(',') if tag.strip()]

def main():
    workspace_dir = "/Users/maurolollo/Desktop/Academy Library"
    cms_path = os.path.join(workspace_dir, "Academy CMS Master 1.xlsx")
    track_path = os.path.join(workspace_dir, "Academy Track Master 1.xlsx")
    
    # 1. Parse Assets
    print("Parsing assets sheet...")
    wb_cms = openpyxl.load_workbook(cms_path, read_only=True)
    sheet_assets = wb_cms['assets']
    
    assets = []
    assets_by_name = {}
    
    # Read headers
    rows_iter = sheet_assets.iter_rows(values_only=True)
    headers = next(rows_iter)
    
    for i, r in enumerate(rows_iter, start=2):
        asset_name = r[1]
        if not asset_name:
            continue
        
        name_clean = asset_name.strip()
        asset_type = clean_string(r[2]) or "video"
        duration_sec = parse_duration(r[3])
        prereq = clean_string(r[4])
        difficulty = r[5]
        if isinstance(difficulty, (int, float)):
            difficulty = float(difficulty)
        else:
            difficulty = None
            
        tags = parse_tags(r[6])
        last_updated = clean_string(r[7])
        cvp_version = clean_string(r[8])
        eos_version = clean_string(r[9])
        avd_version = clean_string(r[10])
        
        needs_update = False
        needs_update_val = clean_string(r[11])
        if needs_update_val and needs_update_val.lower() in ("yes", "true", "1"):
            needs_update = True
            
        comments = clean_string(r[12])
        topic = clean_string(r[13])
        
        asset_id = slugify(name_clean)
        
        asset_doc = {
            "asset_id": asset_id,
            "name": name_clean,
            "type": asset_type,
            "version": 1,
            "is_latest": True,
            "attributes": {
                "duration": duration_sec,
                "prerequisite": prereq,
                "difficulty_level": difficulty,
                "skill_tags": tags,
                "last_updated": last_updated,
                "cvp_version": cvp_version,
                "eos_version": eos_version,
                "avd_version": avd_version,
                "needs_update": needs_update,
                "comments": comments,
                "topic": topic
            }
        }
        
        assets.append(asset_doc)
        assets_by_name[name_clean.lower()] = asset_id
        
    print(f"Loaded {len(assets)} assets.")
    
    # 2. Parse Tracks (Course Elements)
    print("Parsing course elements track structure...")
    wb_track = openpyxl.load_workbook(track_path, read_only=True)
    sheet_course = wb_track['Course Elements']
    
    curriculum = []
    orphaned_count = 0
    
    rows_iter_course = sheet_course.iter_rows(values_only=True)
    course_headers = next(rows_iter_course)
    
    for i, r in enumerate(rows_iter_course, start=2):
        track_name = r[1]
        if not track_name:
            continue
            
        track_name = track_name.strip()
        sub_track_name = clean_string(r[3])
        lesson_name = clean_string(r[5])
        topic_name = clean_string(r[7])
        sub_topic_name = clean_string(r[10]) # Asset Name reference
        
        if not sub_topic_name:
            continue
            
        # Validation: check if asset exists
        asset_key = sub_topic_name.lower()
        if asset_key not in assets_by_name:
            print(f"[ORPHANED ASSET ALERT] Row {i}: Track '{track_name}', Lesson '{lesson_name}', Topic '{topic_name}' references asset '{sub_topic_name}' which is NOT in Assets Master!")
            orphaned_count += 1
            asset_ref_id = None
        else:
            asset_ref_id = assets_by_name[asset_key]
            
        # Capture numeric indicators for sorting metadata
        track_num = r[0] if isinstance(r[0], (int, float)) else None
        sub_track_num = r[2] if isinstance(r[2], (int, float)) else None
        lesson_num = r[4] if isinstance(r[4], (int, float)) else None
        topic_num = r[6] if isinstance(r[6], (int, float)) else None
        sub_topic_num = r[9] if isinstance(r[9], (int, float)) else None
        
        curriculum_item = {
            "track_id": slugify(track_name),
            "track_name": track_name,
            "sub_track": sub_track_name,
            "lesson": lesson_name,
            "topic": topic_name,
            "asset_name": sub_topic_name,
            "asset_ref_id": asset_ref_id,
            "version": 1,
            "is_latest": True,
            "sorting": {
                "track_number": track_num,
                "sub_track_number": sub_track_num,
                "lesson_number": lesson_num,
                "topic_number": topic_num,
                "sub_topic_number": sub_topic_num
            }
        }
        
        curriculum.append(curriculum_item)
        
    print(f"Loaded {len(curriculum)} curriculum items. Orphaned count: {orphaned_count}")
    
    # Save output JSON
    etl_dir = os.path.join(workspace_dir, "etl")
    os.makedirs(etl_dir, exist_ok=True)
    out_path = os.path.join(etl_dir, "data.json")
    
    with open(out_path, "w") as f:
        json.dump({"assets": assets, "curriculum": curriculum}, f, indent=2)
        
    print(f"Data successfully extracted and written to {out_path}")

if __name__ == "__main__":
    main()
