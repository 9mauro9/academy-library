import sys
import os
import json
import urllib.request
import urllib.error
import openpyxl

def parse_excel_to_text(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    structured_data = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = []
        for r in sheet.iter_rows(values_only=True):
            # Clean row values to avoid None clutter
            cleaned_row = [str(x).strip() if x is not None else "" for x in r]
            # Ignore completely empty rows
            if any(cleaned_row):
                rows.append(cleaned_row)
        
        # Limit rows to avoid huge prompt sizes (e.g. first 200 rows per sheet)
        structured_data[sheet_name] = rows[:200]
        
    return structured_data

def call_gemini(api_key, data):
    # Formulate prompt instruction
    prompt = (
        "You are an expert data migration AI. You are given unstructured Excel sheets containing "
        "training curriculum tracks, topics, lessons, and content assets. "
        "Analyze this structure and extract: \n"
        "1. Assets (the unique content items like videos, quizzes, labs, documents).\n"
        "2. Curriculum map mappings (how these assets are ordered into Track -> Sub-Track -> Lesson -> Topic -> Sub-Topic).\n\n"
        "Here is the raw sheet data as a JSON object where keys are sheet names and values are arrays of rows:\n"
        f"{json.dumps(data, indent=2)}\n\n"
        "INSTRUCTIONS:\n"
        "- Group the curriculum structure logically. A Track can be named after the sheet name or header cells.\n"
        "- Sub-topics map to individual assets. The sub_topic_name in the curriculum map MUST match the asset name exactly.\n"
        "- Detect asset types: if name contains 'quiz' it is a 'quiz'; if it represents hands-on or CLI work, it is a 'lab'; if it is text/slides, it is a 'document'; otherwise default to 'video'.\n"
        "- Re-format durations: if length is represented as minutes or time delta, convert it to integer seconds.\n"
        "- Output should conform EXACTLY to this JSON structure:\n"
        "{\n"
        '  "assets": [\n'
        "    {\n"
        '      "asset_id": "alphanumeric-unique-id",\n'
        '      "name": "Asset Name",\n'
        '      "type": "video|lab|quiz|document",\n'
        '      "version": 1,\n'
        '      "attributes": {\n'
        '        "duration": 300,  // seconds\n'
        '        "difficulty_level": 3.5,  // float 1.0 to 5.0\n'
        '        "skill_tags": ["tag1", "tag2"],\n'
        '        "topic": "Topic Name",\n'
        '        "cvp_version": "version",\n'
        '        "eos_version": "version",\n'
        '        "prerequisite": "prereq-asset-id",\n'
        '        "needs_update": false,\n'
        '        "comments": "any comments"\n'
        "      }\n"
        "    }\n"
        "  ],\n"
        '  "curriculum": [\n'
        "    {\n"
        '      "track_name": "Track Name",\n'
        '      "sub_track_name": "Sub Track Name",\n'
        '      "lesson_name": "Lesson Name",\n'
        '      "topic_name": "Topic Name",\n'
        '      "sub_topic_name": "Asset Name",  // MUST match asset name above exactly\n'
        '      "duration": 300,\n'
        '      "difficulty_level": 3.5,\n'
        '      "skill_tag": "tag1, tag2",\n'
        '      "description": "lesson/topic description"\n'
        "    }\n"
        "  ]\n"
        "}\n"
        "Return ONLY the valid JSON object fitting this schema."
    )

    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
    payload = {
        "contents": [{
            "parts": [{
                "text": prompt
            }]
        }],
        "generationConfig": {
            "responseMimeType": "application/json"
        }
    }

    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode('utf-8'),
        headers={'Content-Type': 'application/json'},
        method='POST'
    )

    try:
        with urllib.request.urlopen(req) as response:
            res_data = response.read().decode('utf-8')
            res_json = json.loads(res_data)
            # Extract text from response structure
            text_content = res_json['candidates'][0]['content']['parts'][0]['text']
            return text_content
    except urllib.error.HTTPError as e:
        print(f"HTTP Error: {e.code} - {e.read().decode('utf-8')}", file=sys.stderr)
        raise e
    except Exception as e:
        print(f"Error calling Gemini: {e}", file=sys.stderr)
        raise e

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 gemini_parse_excel.py <filepath>", file=sys.stderr)
        sys.exit(1)

    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}", file=sys.stderr)
        sys.exit(1)

    # 1. Load Gemini API Key
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        # Check local .env file manually
        env_path = os.path.join(os.path.dirname(__file__), "..", ".env")
        if os.path.exists(env_path):
            with open(env_path, "r") as f:
                for line in f:
                    if line.strip().startswith("GEMINI_API_KEY="):
                        api_key = line.strip().split("=", 1)[1].strip()
                        # strip potential quotes
                        if api_key.startswith('"') and api_key.endswith('"'):
                            api_key = api_key[1:-1]
                        if api_key.startswith("'") and api_key.endswith("'"):
                            api_key = api_key[1:-1]
                        break

    if not api_key:
        print("Error: GEMINI_API_KEY environment variable is not defined.", file=sys.stderr)
        sys.exit(1)

    try:
        # 2. Parse Excel
        print(f"Parsing Excel sheets from {file_path}...", file=sys.stderr)
        excel_data = parse_excel_to_text(file_path)

        # 3. Call Gemini
        print("Sending unstructured data to Gemini AI for extraction...", file=sys.stderr)
        json_output = call_gemini(api_key, excel_data)

        # 4. Output the extracted JSON
        print(json_output)
        sys.exit(0)
    except Exception as e:
        print(f"Execution failed: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()
